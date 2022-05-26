from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl


driverEdge = webdriver.Edge()
driverEdge.get('https://www.google.com.au/')
#driverEdge.implicitly_wait(10)
wait = WebDriverWait(driverEdge, 10)
driverEdge.maximize_window()
wb = openpyxl.load_workbook('allEmployers_1.xlsx')
ws = wb.active

num = len(ws['A'])
for i in range(6188, num+1):
    employerName = ws.cell(row=i, column=1).value + ' au'
    # print(employerName)
    xpath_value = '//input[@class="gLFyf gsfi"]'
    # driverEdge.find_element(By.XPATH, '//input[@class="gLFyf gsfi"]').send_keys(employerName)
    # driverEdge.find_element(By.XPATH, '//div[@class="CqAVzb lJ9FBc"]//input[@class="gNO89b"]').click()
    wait.until(EC.presence_of_element_located((By.XPATH, xpath_value))).send_keys(employerName)
    wait.until(EC.presence_of_element_located((By.XPATH, xpath_value))).send_keys(Keys.RETURN)

    xpath_address = '//span[@class="LrzXr"]'
    xpath_phone = '//span[@class="LrzXr zdqRlf kno-fv"]//span/span'

    if len(driverEdge.find_elements(By.XPATH, xpath_phone)) > 0:
        phone = wait.until(EC.presence_of_element_located((By.XPATH, xpath_phone))).text
        ws.cell(row=i, column=2).value = phone

    if len(driverEdge.find_elements(By.XPATH, xpath_address)) > 0:
        address = wait.until(EC.presence_of_element_located((By.XPATH, xpath_address))).text
        ws.cell(row=i, column=3).value = address

    wb.save('allEmployers_1.xlsx')
    # clear the input box
    driverEdge.find_element(By.XPATH, '//div[@class="BKRPef M2vV3"]').click()

driverEdge.quit()
